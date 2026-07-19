import io
import re
import shutil
from base64 import b64encode
from dataclasses import dataclass
from datetime import datetime, time
from html import escape
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
import qrcode
import streamlit as st


st.set_page_config(
    page_title="Menu Digital - Restaurant La Galette",
    layout="wide",
)


@dataclass(frozen=True)
class Dish:
    row_number: int
    name: str
    category: str
    description: str
    composition: str
    price_label: str
    price_value: int
    image_value: str
    image_path: Path | None
    available: bool
    service: str


BASE_DIR = Path(__file__).parent
MENU_FILE = BASE_DIR / "Menu la Galette.xlsx"
IMAGE_DIR = BASE_DIR / "image"
SHEET_NAME = "Restaurant"
SERVICE_OPTIONS = ["Déjeuner", "Dîner", "Déjeuner et dîner"]
LUNCH_START = time(8, 0)
LUNCH_END = time(17, 0)
MENU_HEADERS = [
    "Plat",
    "Description",
    "Catégorie",
    "Composition",
    "image",
    "Prix (FCFA)",
    "Disponible",
    "Service",
]


def get_admin_password() -> str:
    try:
        return st.secrets["ADMIN_PASSWORD"]
    except Exception:
        return "admin123"


def clean_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def is_available(value: object) -> bool:
    text = clean_text(value).lower()
    return text not in {"non", "no", "false", "0", "indisponible"}


def available_label(value: bool) -> str:
    return "Oui" if value else "Non"


def normalize_service(value: object, dish_name: str = "") -> str:
    text = clean_text(value)
    if text in SERVICE_OPTIONS:
        return text

    lowered = text.lower()
    if lowered in {"dejeuner", "déjeuner", "midi", "lunch"}:
        return "Déjeuner"
    if lowered in {"diner", "dîner", "soir", "dinner"}:
        return "Dîner"
    if lowered in {"les deux", "dejeuner et diner", "déjeuner et dîner", "tous", "tout"}:
        return "Déjeuner et dîner"

    name = dish_name.lower()
    if "thiébou" in name or "thiebou" in name or "ceebu" in name:
        return "Déjeuner"
    return "Déjeuner et dîner"


def service_matches_filter(dish_service: str, selected_service: str) -> bool:
    if selected_service == "Tout":
        return True
    return dish_service == selected_service or dish_service == "Déjeuner et dîner"


def current_service(now: datetime | None = None) -> str:
    current_time = (now or datetime.now()).time()
    if LUNCH_START <= current_time < LUNCH_END:
        return "Déjeuner"
    return "Dîner"


def parse_price_value(value: object) -> int:
    numbers = [int(match.replace(" ", "")) for match in re.findall(r"\d[\d ]*", clean_text(value))]
    return min(numbers) if numbers else 0


def format_price(value: object) -> str:
    text = clean_text(value)
    if not text:
        return "Prix non indiqué"
    return text if "FCFA" in text.upper() else f"{text} FCFA"


def normalize_image_path(value: str) -> Path | None:
    if not value:
        return None
    return BASE_DIR / value.replace("\\", "/")


def load_sheet():
    try:
        workbook = load_workbook(MENU_FILE)
        sheet = workbook[SHEET_NAME]
        return workbook, sheet
    except PermissionError as error:
        raise RuntimeError("Fermez le fichier Excel avant d'enregistrer une modification.") from error
    except BadZipFile as error:
        raise RuntimeError("Le fichier Excel semble ouvert ou endommagé. Fermez Excel puis réessayez.") from error


def header_indexes(sheet) -> dict[str, int]:
    headers = [clean_text(cell.value) for cell in sheet[1]]
    return {header: index + 1 for index, header in enumerate(headers) if header}


def ensure_menu_schema() -> None:
    workbook, sheet = load_sheet()
    indexes = header_indexes(sheet)
    changed = False

    for header in MENU_HEADERS:
        if header not in indexes:
            sheet.cell(row=1, column=sheet.max_column + 1, value=header)
            changed = True
            indexes = header_indexes(sheet)

    if changed:
        available_col = indexes["Disponible"]
        service_col = indexes["Service"]
        name_col = indexes["Plat"]
        for row_number in range(2, sheet.max_row + 1):
            if clean_text(sheet.cell(row=row_number, column=available_col).value) == "":
                sheet.cell(row=row_number, column=available_col, value="Oui")
            if clean_text(sheet.cell(row=row_number, column=service_col).value) == "":
                dish_name = clean_text(sheet.cell(row=row_number, column=name_col).value)
                sheet.cell(row=row_number, column=service_col, value=normalize_service("", dish_name))
        workbook.save(MENU_FILE)


@st.cache_data(show_spinner=False)
def load_menu(include_unavailable: bool = False) -> list[Dish]:
    workbook = load_workbook(MENU_FILE, data_only=True)
    sheet = workbook[SHEET_NAME]
    indexes = header_indexes(sheet)
    dishes = []

    for row_number in range(2, sheet.max_row + 1):
        def value(header: str) -> object:
            column = indexes.get(header)
            return sheet.cell(row=row_number, column=column).value if column else ""

        name = clean_text(value("Plat"))
        if not name:
            continue

        image_value = clean_text(value("image"))
        image_path = normalize_image_path(image_value)
        available = is_available(value("Disponible"))
        service = normalize_service(value("Service"), name)

        if not include_unavailable and not available:
            continue

        dishes.append(
            Dish(
                row_number=row_number,
                name=name,
                description=clean_text(value("Description")),
                category=clean_text(value("Catégorie")) or "Autres",
                composition=clean_text(value("Composition")),
                price_label=format_price(value("Prix (FCFA)")),
                price_value=parse_price_value(value("Prix (FCFA)")),
                image_value=image_value,
                image_path=image_path if image_path and image_path.exists() else None,
                available=available,
                service=service,
            )
        )

    return dishes


def refresh_menu_cache() -> None:
    load_menu.clear()
    image_to_data_uri.clear()


def update_row(row_number: int, updates: dict[str, object]) -> None:
    ensure_menu_schema()
    workbook, sheet = load_sheet()
    indexes = header_indexes(sheet)

    for header, value in updates.items():
        sheet.cell(row=row_number, column=indexes[header], value=value)

    try:
        workbook.save(MENU_FILE)
    except PermissionError as error:
        raise RuntimeError("Fermez le fichier Excel avant d'enregistrer une modification.") from error
    refresh_menu_cache()


def append_dish(values: dict[str, object]) -> None:
    ensure_menu_schema()
    workbook, sheet = load_sheet()
    indexes = header_indexes(sheet)
    row_number = sheet.max_row + 1

    for header in MENU_HEADERS:
        sheet.cell(row=row_number, column=indexes[header], value=values.get(header, ""))

    try:
        workbook.save(MENU_FILE)
    except PermissionError as error:
        raise RuntimeError("Fermez le fichier Excel avant d'ajouter un plat.") from error
    refresh_menu_cache()


def save_uploaded_image(uploaded_file) -> str:
    IMAGE_DIR.mkdir(exist_ok=True)
    safe_name = re.sub(r"[^A-Za-z0-9_. -]", "", uploaded_file.name).strip() or "plat.jpg"
    destination = IMAGE_DIR / safe_name
    counter = 1

    while destination.exists():
        destination = IMAGE_DIR / f"{destination.stem}-{counter}{destination.suffix}"
        counter += 1

    with destination.open("wb") as output:
        shutil.copyfileobj(uploaded_file, output)

    return f"image\\{destination.name}"


@st.cache_data(show_spinner=False)
def image_to_data_uri(path: Path | None) -> str:
    if path is None or not path.exists():
        return ""

    mime = "image/jpeg" if path.suffix.lower() in {".jpg", ".jpeg", ".jfif"} else "image/png"
    encoded = b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def build_qr_code(url: str) -> bytes:
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=12,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#111111", back_color="#ffffff")
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    return buffer.getvalue()


def inject_styles() -> None:
    st.markdown(
        """
        <style>
        .stApp {
            background: #fbfaf7;
            color: #1f1f1d;
        }
        section[data-testid="stSidebar"] {
            background: #f1eee8;
        }
        .hero {
            padding: 28px 0 16px;
            border-bottom: 1px solid #ded8cc;
            margin-bottom: 22px;
        }
        .hero h1 {
            font-size: 42px;
            line-height: 1.08;
            margin: 0 0 8px;
            letter-spacing: 0;
        }
        .hero p {
            max-width: 760px;
            font-size: 17px;
            color: #615d55;
            margin: 0;
        }
        .dish-card {
            background: #ffffff;
            border: 1px solid #e4ded3;
            border-radius: 8px;
            overflow: hidden;
            margin-bottom: 22px;
            box-shadow: 0 8px 20px rgba(50, 45, 37, 0.06);
        }
        .dish-image {
            width: 100%;
            height: 220px;
            object-fit: cover;
            display: block;
            background: #ede7dc;
        }
        .dish-body {
            padding: 16px 16px 18px;
        }
        .dish-topline {
            display: flex;
            justify-content: space-between;
            gap: 12px;
            align-items: flex-start;
        }
        .dish-title {
            font-size: 20px;
            font-weight: 700;
            margin: 0;
        }
        .dish-price {
            white-space: nowrap;
            font-weight: 800;
            color: #9a3f22;
        }
        .dish-description {
            color: #5f5a51;
            margin-top: 9px;
            min-height: 44px;
        }
        .dish-composition {
            color: #756f65;
            font-size: 14px;
            margin-top: 10px;
            padding-top: 10px;
            border-top: 1px solid #eee8de;
        }
        .dish-category {
            display: inline-flex;
            margin-top: 13px;
            color: #633513;
            font-size: 13px;
            font-weight: 800;
        }
        .dish-service {
            display: inline-flex;
            margin-top: 10px;
            margin-right: 8px;
            padding: 4px 9px;
            border-radius: 999px;
            background: #f2dfc3;
            color: #633513;
            font-size: 12px;
            font-weight: 800;
        }
        .section-title {
            font-size: 24px;
            margin: 8px 0 18px;
            font-weight: 800;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_dish_card(dish: Dish) -> None:
    image_src = image_to_data_uri(dish.image_path)
    image_html = (
        f'<img class="dish-image" src="{image_src}" alt="{escape(dish.name)}">'
        if image_src
        else '<div class="dish-image"></div>'
    )
    st.markdown(
        f"""
        <article class="dish-card">
            {image_html}
            <div class="dish-body">
                <div class="dish-topline">
                    <h3 class="dish-title">{escape(dish.name)}</h3>
                    <div class="dish-price">{escape(dish.price_label)}</div>
                </div>
                <div class="dish-description">{escape(dish.description)}</div>
                <div class="dish-composition"><strong>Composition :</strong> {escape(dish.composition)}</div>
                <div class="dish-service">{escape(dish.service)}</div>
                <div class="dish-category">{escape(dish.category)}</div>
            </div>
        </article>
        """,
        unsafe_allow_html=True,
    )


def render_menu_page() -> None:
    menu = load_menu()
    automatic_service = current_service()
    st.markdown(
        """
        <div class="hero">
            <h1>Restaurant La Galette</h1>
            <p>Découvrez nos plats, grillades, pâtes, spécialités de la mer et recettes traditionnelles.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    service_mode = st.sidebar.radio(
        "Menu temporel",
        ["Automatique", "Choisir manuellement"],
        index=0,
    )
    if service_mode == "Automatique":
        selected_service = automatic_service
        st.sidebar.caption(f"Service actuel : {automatic_service} (08h00-17h00 = déjeuner)")
    else:
        selected_service = st.sidebar.radio("Service", ["Tout", *SERVICE_OPTIONS], index=0)

    # 1. Filtrer d'abord les plats selon le service choisi
    service_dishes = [
        dish
        for dish in menu
        if service_matches_filter(dish.service, selected_service)
    ]

    # 2. Récupérer et sélectionner la catégorie correspondante
    categories = ["Tout"] + sorted({dish.category for dish in service_dishes})
    selected_category = st.sidebar.radio("Catégorie", categories, index=0)

    # 3. Filtrer par catégorie AVANT d'analyser le prix max
    category_dishes = [
        dish 
        for dish in service_dishes 
        if selected_category == "Tout" or dish.category == selected_category
    ]

    # 4. Calculer le prix maximum disponible dans cette sélection de plats
    highest_price = max((dish.price_value for dish in category_dishes), default=0)

    # 5. Rendre le slider de prix de façon ultra sécurisée
    if highest_price <= 0:
        max_price = st.sidebar.slider(
            "Prix maximum",
            min_value=0,
            max_value=500,  # Empêche la plage nulle (0-0)
            value=0,
            step=500,
        )
    else:
        max_price = st.sidebar.slider(
            "Prix maximum",
            min_value=0,
            max_value=highest_price,
            value=highest_price,
            step=500,
        )

    # 6. Recherche par mots-clés
    search = st.sidebar.text_input("Rechercher un plat", placeholder="Ex: poulet, gambas...")

    # 7. Appliquer le filtre final de prix et de recherche textuelle
    visible_dishes = [
        dish
        for dish in category_dishes
        if dish.price_value <= max_price
        and (
            not search
            or search.lower() in dish.name.lower()
            or search.lower() in dish.description.lower()
            or search.lower() in dish.composition.lower()
        )
    ]

    st.markdown(
        f'<div class="section-title">{len(visible_dishes)} option(s) disponible(s) - {escape(selected_service)}</div>',
        unsafe_allow_html=True,
    )

    if not visible_dishes:
        st.info("Aucun plat ne correspond à ces filtres.")
        return

    columns = st.columns(3)
    for index, dish in enumerate(visible_dishes):
        with columns[index % 3]:
            render_dish_card(dish)


def render_qr_page() -> None:
    st.markdown(
        """
        <div class="hero">
            <h1>QR code du menu</h1>
            <p>Déployez l'application, collez son lien public ici, puis téléchargez le QR code à imprimer.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    default_url = "https://votre-menu.streamlit.app"
    menu_url = st.text_input("Lien public du menu", value=default_url)

    if menu_url:
        qr_bytes = build_qr_code(menu_url)
        st.image(qr_bytes, caption=menu_url, width=320)
        st.download_button(
            "Télécharger le QR code",
            data=qr_bytes,
            file_name="qr-code-menu-restaurant.png",
            mime="image/png",
        )

    st.info(
        "Pour l'utiliser avec vos clients, hébergez l'application sur Streamlit Community Cloud, Render, ou un serveur web. "
        "Le QR code doit pointer vers l'adresse publique obtenue après le déploiement."
    )


def render_admin_login() -> bool:
    if st.session_state.get("admin_authenticated"):
        return True

    st.markdown(
        """
        <div class="hero">
            <h1>Administration</h1>
            <p>Connectez-vous pour gérer les disponibilités, les prix et les nouveaux plats.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.form("admin_login_form"):
        password = st.text_input("Mot de passe administrateur", type="password")
        submitted = st.form_submit_button("Se connecter")

    if submitted:
        if password == get_admin_password():
            st.session_state["admin_authenticated"] = True
            st.rerun()
        else:
            st.error("Mot de passe incorrect.")

    st.caption("Mot de passe par défaut en local : admin123. Changez-le avant un déploiement public.")
    return False


def render_availability_admin(dishes: list[Dish]) -> None:
    st.subheader("Disponibilité et prix")
    st.caption("Décochez un plat indisponible : il disparaîtra du menu client.")

    for dish in dishes:
        with st.expander(f"{dish.name} - {dish.price_label}", expanded=False):
            col_status, col_price = st.columns([1, 2])
            available = col_status.checkbox(
                "Disponible",
                value=dish.available,
                key=f"available_{dish.row_number}",
            )
            price = col_price.text_input(
                "Prix affiché",
                value=dish.price_label.replace(" FCFA", ""),
                key=f"price_{dish.row_number}",
            )
            service = st.selectbox(
                "Service",
                SERVICE_OPTIONS,
                index=SERVICE_OPTIONS.index(dish.service),
                key=f"service_{dish.row_number}",
            )

            if st.button("Enregistrer", key=f"save_quick_{dish.row_number}"):
                try:
                    update_row(
                        dish.row_number,
                        {
                            "Disponible": available_label(available),
                            "Prix (FCFA)": price,
                            "Service": service,
                        },
                    )
                except RuntimeError as error:
                    st.error(str(error))
                else:
                    st.success("Plat mis à jour.")
                    st.rerun()


def render_edit_admin(dishes: list[Dish]) -> None:
    st.subheader("Modifier un plat")
    dish_names = [dish.name for dish in dishes]

    if not dish_names:
        st.info("Aucun plat à modifier.")
        return

    selected_name = st.selectbox("Plat à modifier", dish_names)
    dish = next(item for item in dishes if item.name == selected_name)

    with st.form("edit_dish_form"):
        name = st.text_input("Nom du plat", value=dish.name, key="edit_name")
        description = st.text_area("Description", value=dish.description, key="edit_description")
        category = st.text_input("Catégorie", value=dish.category, key="edit_category")
        composition = st.text_area("Composition", value=dish.composition, key="edit_composition")
        price = st.text_input("Prix (FCFA)", value=dish.price_label.replace(" FCFA", ""), key="edit_price")
        service = st.selectbox(
            "Service",
            SERVICE_OPTIONS,
            index=SERVICE_OPTIONS.index(dish.service),
            key="edit_service",
        )
        available = st.checkbox("Disponible", value=dish.available, key="edit_available")
        image_value = st.text_input("Chemin image", value=dish.image_value, key="edit_image_value")
        uploaded_file = st.file_uploader(
            "Remplacer l'image",
            type=["jpg", "jpeg", "png", "jfif"],
            key="edit_uploaded_file",
        )
        submitted = st.form_submit_button("Enregistrer les modifications")

    if submitted:
        try:
            if uploaded_file is not None:
                image_value = save_uploaded_image(uploaded_file)

            update_row(
                dish.row_number,
                {
                    "Plat": name,
                    "Description": description,
                    "Catégorie": category,
                    "Composition": composition,
                    "image": image_value,
                    "Prix (FCFA)": price,
                    "Disponible": available_label(available),
                    "Service": service,
                },
            )
        except RuntimeError as error:
            st.error(str(error))
        else:
            st.success("Plat modified.")
            st.rerun()


def render_add_admin() -> None:
    st.subheader("Ajouter un nouveau plat")

    with st.form("add_dish_form"):
        name = st.text_input("Nom du plat", key="add_name")
        description = st.text_area("Description", key="add_description")
        category = st.text_input("Catégorie", key="add_category")
        composition = st.text_area("Composition", key="add_composition")
        price = st.text_input("Prix (FCFA)", placeholder="Ex: 3 500", key="add_price")
        service = st.selectbox("Service", SERVICE_OPTIONS, index=2, key="add_service")
        available = st.checkbox("Disponible", value=True, key="add_available")
        uploaded_file = st.file_uploader(
            "Image du plat",
            type=["jpg", "jpeg", "png", "jfif"],
            key="add_uploaded_file",
        )
        submitted = st.form_submit_button("Ajouter le plat")

    if submitted:
        if not name.strip():
            st.error("Le nom du plat est obligatoire.")
            return
        if not price.strip():
            st.error("Le prix est obligatoire.")
            return

        try:
            image_value = save_uploaded_image(uploaded_file) if uploaded_file is not None else ""
            append_dish(
                {
                    "Plat": name.strip(),
                    "Description": description.strip(),
                    "Catégorie": category.strip() or "Autres",
                    "Composition": composition.strip(),
                    "image": image_value,
                    "Prix (FCFA)": price.strip(),
                    "Disponible": available_label(available),
                    "Service": service,
                }
            )
        except RuntimeError as error:
            st.error(str(error))
        else:
            st.success("Nouveau plat ajouté.")
            st.rerun()


def render_admin_page() -> None:
    ensure_menu_schema()
    if not render_admin_login():
        return

    st.markdown(
        """
        <div class="hero">
            <h1>Administration</h1>
            <p>Gérez le menu sans toucher directement au fichier Excel.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if st.button("Se déconnecter"):
        st.session_state.pop("admin_authenticated", None)
        st.rerun()

    dishes = load_menu(include_unavailable=True)
    tab_availability, tab_add, tab_edit = st.tabs(["Disponibilité & prix", "Ajouter", "Modifier"])

    with tab_availability:
        render_availability_admin(dishes)
    with tab_add:
        render_add_admin()
    with tab_edit:
        render_edit_admin(dishes)


def main() -> None:
    inject_styles()

    page = st.sidebar.selectbox("Navigation", ["Menu client", "QR code", "Administration"])
    st.sidebar.caption("Les plats viennent du fichier Menu la Galette.xlsx.")

    if page == "Menu client":
        render_menu_page()
    elif page == "QR code":
        render_qr_page()
    else:
        render_admin_page()


if __name__ == "__main__":
    main()
